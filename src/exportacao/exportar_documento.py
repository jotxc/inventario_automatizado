import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill


def exportar_documento(
    estoque,
    nome_arquivo
):

    colunas_manter = {
        "material": "Material",
        "descricao_material": "Descrição",
        "lote": "Lote",
        "posicao": "Posição",
        "tipo_deposito": "Tipo Depósito",
        "estoque_total": "Qtd. Estoque",
    }

    estoque = estoque[list(colunas_manter.keys())].rename(columns=colunas_manter)

    caminho = (
        Path(__file__)
        .resolve()
        .parents[2]
        / "data"
        / "saida"
    )

    caminho.mkdir(
        parents=True,
        exist_ok=True
    )

    arquivo = caminho / nome_arquivo

    estoque.to_excel(
        arquivo,
        index=False
    )

    workbook = load_workbook(arquivo)

    worksheet = workbook.active

    # Cabeçalho

    for celula in worksheet[1]:
        celula.font = Font(bold=True, color="FFFFFF")
        celula.fill = PatternFill("solid", fgColor="002060")
        celula.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

    # Dados

    for linha in worksheet.iter_rows(min_row=2):
        for celula in linha:
            celula.alignment = Alignment(
                horizontal="left",
                vertical="center"
            )

    # Filtros

    worksheet.auto_filter.ref = worksheet.dimensions

    # Congelar primeira linha

    worksheet.freeze_panes = "A2"

    # Ajustar largura das colunas

    for coluna in worksheet.columns:
        tamanho = max(
            len(str(c.value))
            if c.value is not None
            else 0
            for c in coluna
        )
        worksheet.column_dimensions[
            coluna[0].column_letter
        ].width = tamanho + 4

    print(f"\nArquivo será salvo em:\n{arquivo}\n")

    workbook.save(arquivo)